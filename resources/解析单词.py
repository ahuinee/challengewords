import openpyxl #xlsx解析
import re # 正则
import json # JSON
# 文件名
fileName = "单词.xlsx"
# 以只读模式打开工作簿
wb = openpyxl.load_workbook(fileName, read_only=True)

# 获取workbook中所有的表格sheet
sheets = wb.get_sheet_names()

# 班级与书册
volumeRE = re.compile("(.*)(上册|下册)")
# 单元匹配
unitRE = re.compile(".*单元")
# 单词匹配 
wordRE_EN = re.compile("[a-zA-Z]")

# 年级
gradeList = []
# 书册
volumeList = []
# 循环遍历所有sheet len(sheets)
for i in range(len(sheets)):
    # 每个sheet 年级和书册
    sheet = wb.get_sheet_by_name(sheets[i])
    # 年级与书册
    st = volumeRE.match(sheet.title)
    if st:
        gradeName = st.group(1) # 年级
        volumeName = st.group(2) # 书册
        volumeDict = {'volumeName':volumeName,'units':[]}
        # 年级
        if volumeName == '上册':
            gradeDict = {'gradeName':gradeName,'volumes':[]}
            volumeList = gradeDict['volumes']
            gradeList.append(gradeDict)
        # 书册
        volumeList.append(volumeDict)
        # 单元数组
        unitList = volumeDict['units']
        unitDict = None
        for c in range(1, sheet.max_column+1):
            for r in range(1,sheet.max_row+1):
                # 每个cell的内容
                cellValue = sheet.cell(r, c).value
                if cellValue :
                    cellValue = str(cellValue)
                    # 单元
                    if unitRE.match(cellValue):
                        unit = {
                            "unitName":cellValue,
                            "words": []
                        }
                        unitDict = unit
                        unitList.append(unit)
                        continue
                    # 过滤掉English
                    if wordRE_EN.match(cellValue):
                        continue
                    # 单词 (仅中文)
                    if unitDict :
                        word_en = str(sheet.cell(r, c+1).value) # 取中文对应掉英语单词
                        unitDict['words'].append({'chinese':cellValue,'english':word_en})

    
# 将json对象写入到文件
with open('words.json', 'w') as dump_f:
    json.dump(gradeList,dump_f,indent=4,ensure_ascii=False)
    dump_f.close()
